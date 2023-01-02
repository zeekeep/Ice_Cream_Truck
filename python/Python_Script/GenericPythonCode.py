import os
import csv
import glob
import time
import shutil
from O365 import Account
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.workbook import Workbook

credentials = (
    "5f2bf0de-5d68-4c0a-833f-b32899de4579",
    "Xmv8Q~Kdk4wHu4ARgHOMRx8CVRwPZ.4gqe_4JcRD",
)
account = Account(
    credentials,
    auth_flow_type="credentials",
    tenant_id="d4295056-baca-4708-8c60-0351c0fb01db",
)

if account.authenticate():
    driver = webdriver.Firefox(
        executable_path="/home/abhishekraj/python/Python_Script/geckodriver"
    )
    driver.get(
        "https://security.microsoft.com/securitypoliciesandrules?tid=d4295056-baca-4708-8c60-0351c0fb01db"
    )
    driver.maximize_window()
    email = driver.find_element(By.NAME, "loginfmt")
    time.sleep(4)
    email.send_keys("Abhishek@jeeshanahmad2011outlook.onmicrosoft.com")
    time.sleep(4)
    driver.find_element(By.ID, "idSIButton9").click()
    pswd = driver.find_element(By.NAME, "passwd")
    pswd.send_keys("12345@Fdsa")
    time.sleep(4)
    signIn = driver.find_element(By.XPATH, "//input[@id='idSIButton9']").click()
    time.sleep(4)
    loginIn = driver.find_element(By.ID, "idSIButton9").click()
    time.sleep(10)
    back_to_all_policy = driver.find_element(
        By.XPATH, "//a[normalize-space()='Threat policies']"
    ).click()
    time.sleep(4)
    ############firstPolicy#############
    policy_number_first = driver.find_element(
        By.XPATH, "//button[normalize-space()='Anti-phishing']"
    ).click()
    time.sleep(4)
    policy_number_first_data = driver.find_elements(
        By.CLASS_NAME, "ms-DetailsRow-fields"
    )
    wb = Workbook()
    first_filepath="/home/abhishekraj/python/Python_Script/all_policy/Anti-phishing.xlsx"
    for data in policy_number_first_data:
        data.click()
        time.sleep(4)
        policy_first_data = driver.find_elements(
            By.CLASS_NAME, "ms-MetaDataItem"
        )
        ws1 = wb.create_sheet(data.text.splitlines()[0])
        for data in policy_first_data:
            sav_data = [data.text.splitlines()]
            for row in sav_data:
                ws1.append(row)
        time.sleep(4)
        close_btn = driver.find_element(
                By.XPATH, "//span[contains(text(),'Close')]"
            ).click()
    wb.save(filename = first_filepath)
    time.sleep(4)
    driver.back()
    time.sleep(4)
    ############secondPolicy#############
    inside_second_Policy_Folder = driver.find_element(
        By.XPATH, "//button[normalize-space()='Anti-spam']"
        ).click()
    time.sleep(4)
    second_Policy_Item = driver.find_elements(By.CLASS_NAME, "ms-DetailsRow-fields")
    time.sleep(4)
    wb = Workbook()
    second_filepath="/home/abhishekraj/python/Python_Script/all_policy/Anti-spam.xlsx"
    for data in second_Policy_Item:
        ws1 = wb.create_sheet(data.text.splitlines()[0])
        data.click()
        time.sleep(4)
        all_items = driver.find_elements(By.CLASS_NAME, "ms-MetaDataItem")
        semi_final = []
        for res in all_items:
            second_policy_result = res.text.splitlines()
            semi_final.append(second_policy_result)
            sav_data = [res.text.splitlines()]
            for row in sav_data:
                ws1.append(row)
        wb.save(filename = second_filepath)
        time.sleep(3)
        try:
            button = driver.find_element(
                By.XPATH,
                "//button[@aria-label='Edit allowed and blocked senders and domains']",
            )
            time.sleep(4)
            if button.is_displayed():
                driver.execute_script("arguments[0].click();", button)
                get_data = driver.find_elements(
                    By.XPATH,
                    "//div[@id='fluent-default-layer-host']/div/div/div/div/div[2]/div[2]/div/div[4]/div/button",
                )
                time.sleep(4)
                for i in get_data:
                    i.click()
                    time.sleep(4)
                    items_ = driver.find_elements(
                        By.CLASS_NAME, "ms-DetailsList-contentWrapper"
                    )
                    for new_res in items_:
                        second_policy_ = new_res.text.splitlines()
                    semi_final.append(second_policy_)
                    save_data = [i.text.splitlines(),second_policy_]
                    a = "\uea3a"
                    try:
                        for a in second_policy_:
                            second_policy_.remove("\uea3a")
                    except:
                        print()
                    flat_list = [item for sublist in save_data for item in sublist]
                    print(flat_list)
                    ws1.append(flat_list)
                    wb.save(filename = second_filepath)
                    semi_final.append(second_policy_)
                    done_btn = driver.find_element(
                        By.XPATH, "//button[@aria-label='Done']"
                    ).click()
                time.sleep(4)
                cancel_btn = driver.find_element(
                    By.XPATH, "//i[@data-icon-name='Cancel']"
                ).click()
            time.sleep(4)
            close_btn = driver.find_element(
                By.XPATH, "//span[contains(text(),'Close')]"
            ).click()
        except:
            time.sleep(4)
            close_btn = driver.find_element(
                By.XPATH, "//span[contains(text(),'Close')]"
            ).click()
    time.sleep(4)
    driver.back()
    time.sleep(4)
    ###########thirdPolicy################
    third_filepath="/home/abhishekraj/python/Python_Script/all_policy/Anti-malware.xlsx"

    policy_third = driver.find_element(
        By.XPATH, "//button[normalize-space()='Anti-malware']"
        ).click()
    time.sleep(4)
    data_policy_third_number_policy = driver.find_elements(
        By.CLASS_NAME, "ms-DetailsRow-fields"
        )
    wb = Workbook()
    for third_data in data_policy_third_number_policy:
        ws1 = wb.create_sheet(third_data.text.splitlines()[0])
        third_data.click()
        time.sleep(4)
        data_policy_third_number = driver.find_elements(
            By.CLASS_NAME, "ms-MetaDataItem"
            )
        for data in data_policy_third_number:
            sav_data = [data.text.splitlines()]
            for row in sav_data:
                ws1.append(row)
            time.sleep(4)
        close_btn = driver.find_element(
            By.XPATH, "//span[contains(text(),'Close')]"
        ).click()
    wb.save(filename = third_filepath)
    driver.back()
    time.sleep(4)
    secure_Score = driver.find_element(
        By.XPATH, "//a[@name='Secure score']"
    ).click()
    time.sleep(4)
    recomm_action = driver.find_element(
        By.XPATH, "//button[@aria-label='Recommended actions']"
    ).click()
    time.sleep(4)
    secure_Score = driver.find_element(
        By.XPATH, "//span[contains(text(),'Export')]"
    ).click()
    time.sleep(4)
    shutil.copy2('/home/abhishekraj/Downloads/Microsoft Secure Score - Microsoft 365 security.csv', '/home/abhishekraj/python/Python_Script/secure_score/Microsoft Secure Score - Microsoft 365 security.csv')
    driver.quit()